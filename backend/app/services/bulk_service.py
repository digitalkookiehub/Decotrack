"""Bulk import/export service for Vendors and Raw Materials via Excel."""
import io
import logging
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models.raw_material import ItemCategory, MaterialUnit, RawMaterial
from app.models.vendor import Vendor

logger = logging.getLogger(__name__)


# ── VENDORS ────────────────────────────────────────────────────

VENDOR_COLUMNS = ["name", "contact_person", "phone", "email", "address", "city", "state", "gstin", "supply_categories", "payment_terms", "notes"]


def export_vendors_excel(db: Session) -> io.BytesIO:
    vendors = db.query(Vendor).filter(Vendor.is_active == True).order_by(Vendor.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"

    # Header row
    headers = ["ID", "Name*", "Contact Person", "Phone", "Email", "Address", "City", "State", "GSTIN", "Supply Categories (comma-sep)", "Payment Terms", "Notes"]
    ws.append(headers)

    # Style header
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    # Data rows
    for v in vendors:
        cats = ", ".join(v.supply_categories) if v.supply_categories else ""
        ws.append([v.id, v.name, v.contact_person, v.phone, v.email, v.address, v.city, v.state, v.gstin, cats, v.payment_terms, v.notes])

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_vendors_excel(db: Session, file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    updated = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows, start=2):
        if not row or not row[1]:  # Name is required (column B)
            continue

        try:
            vendor_id = row[0] if row[0] else None
            name = str(row[1]).strip()
            contact_person = str(row[2]).strip() if row[2] else None
            phone = str(row[3]).strip() if row[3] else None
            email = str(row[4]).strip() if row[4] else None
            address = str(row[5]).strip() if row[5] else None
            city = str(row[6]).strip() if row[6] else None
            state = str(row[7]).strip() if row[7] else None
            gstin = str(row[8]).strip() if row[8] else None
            supply_cats = [s.strip() for s in str(row[9]).split(",")] if row[9] else None
            payment_terms = str(row[10]).strip() if row[10] else None
            notes = str(row[11]).strip() if len(row) > 11 and row[11] else None

            if vendor_id:
                # Update existing
                vendor = db.query(Vendor).filter(Vendor.id == int(vendor_id)).first()
                if vendor:
                    vendor.name = name
                    vendor.contact_person = contact_person
                    vendor.phone = phone
                    vendor.email = email
                    vendor.address = address
                    vendor.city = city
                    vendor.state = state
                    vendor.gstin = gstin
                    vendor.supply_categories = supply_cats
                    vendor.payment_terms = payment_terms
                    vendor.notes = notes
                    updated += 1
                else:
                    errors.append(f"Row {row_num}: Vendor ID {vendor_id} not found")
            else:
                # Create new
                vendor = Vendor(
                    name=name, contact_person=contact_person, phone=phone, email=email,
                    address=address, city=city, state=state, gstin=gstin,
                    supply_categories=supply_cats, payment_terms=payment_terms, notes=notes,
                )
                db.add(vendor)
                created += 1

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    db.commit()
    logger.info("Vendor bulk import: created=%d, updated=%d, errors=%d", created, updated, len(errors))
    return {"created": created, "updated": updated, "errors": errors}


def delete_vendors_bulk(db: Session, vendor_ids: list[int]) -> int:
    count = 0
    for vid in vendor_ids:
        vendor = db.query(Vendor).filter(Vendor.id == vid).first()
        if vendor:
            vendor.is_active = False
            count += 1
    db.commit()
    logger.info("Vendor bulk delete (deactivate): %d vendors", count)
    return count


# ── RAW MATERIALS ──────────────────────────────────────────────

def export_materials_excel(db: Session) -> io.BytesIO:
    materials = (
        db.query(RawMaterial)
        .filter(RawMaterial.is_active == True)
        .order_by(RawMaterial.name)
        .all()
    )
    categories = {c.id: c.name for c in db.query(ItemCategory).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Materials"

    headers = ["ID", "Name*", "SKU*", "Category", "Unit* (PCS/SQM/M/KG/PAIR/LITRE/SET)", "Current Stock", "Reorder Level*", "Last Purchase Rate", "HSN Code", "GST Rate (%)", "Description"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    for m in materials:
        ws.append([
            m.id, m.name, m.sku, categories.get(m.category_id, ""),
            m.unit.value if m.unit else "PCS",
            float(m.current_stock), float(m.reorder_level), float(m.last_purchase_rate),
            m.hsn_code, float(m.gst_rate), m.description,
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_materials_excel(db: Session, file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # Build category lookup
    categories = {c.name.lower(): c.id for c in db.query(ItemCategory).all()}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    updated = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows, start=2):
        if not row or not row[1]:  # Name required
            continue

        try:
            material_id = row[0] if row[0] else None
            name = str(row[1]).strip()
            sku = str(row[2]).strip() if row[2] else None
            category_name = str(row[3]).strip().lower() if row[3] else None
            unit_str = str(row[4]).strip().upper() if row[4] else "PCS"
            current_stock = Decimal(str(row[5])) if row[5] else Decimal("0")
            reorder_level = Decimal(str(row[6])) if row[6] else Decimal("0")
            last_rate = Decimal(str(row[7])) if row[7] else Decimal("0")
            hsn_code = str(row[8]).strip() if row[8] else None
            gst_rate = Decimal(str(row[9])) if row[9] else Decimal("18")
            description = str(row[10]).strip() if len(row) > 10 and row[10] else None

            if not sku:
                errors.append(f"Row {row_num}: SKU is required")
                continue

            # Validate unit
            try:
                unit = MaterialUnit(unit_str)
            except ValueError:
                errors.append(f"Row {row_num}: Invalid unit '{unit_str}'. Use PCS/SQM/M/KG/PAIR/LITRE/SET")
                continue

            category_id = categories.get(category_name) if category_name else None

            if material_id:
                # Update existing
                material = db.query(RawMaterial).filter(RawMaterial.id == int(material_id)).first()
                if material:
                    material.name = name
                    material.sku = sku
                    material.category_id = category_id
                    material.unit = unit
                    material.reorder_level = reorder_level
                    material.last_purchase_rate = last_rate
                    material.hsn_code = hsn_code
                    material.gst_rate = gst_rate
                    material.description = description
                    updated += 1
                else:
                    errors.append(f"Row {row_num}: Material ID {material_id} not found")
            else:
                # Check if SKU exists
                existing = db.query(RawMaterial).filter(RawMaterial.sku == sku).first()
                if existing:
                    errors.append(f"Row {row_num}: SKU '{sku}' already exists (ID: {existing.id})")
                    continue

                material = RawMaterial(
                    name=name, sku=sku, category_id=category_id, unit=unit,
                    current_stock=current_stock, reorder_level=reorder_level,
                    last_purchase_rate=last_rate, hsn_code=hsn_code, gst_rate=gst_rate,
                    description=description,
                )
                db.add(material)
                created += 1

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    db.commit()
    logger.info("Material bulk import: created=%d, updated=%d, errors=%d", created, updated, len(errors))
    return {"created": created, "updated": updated, "errors": errors}


def delete_materials_bulk(db: Session, material_ids: list[int]) -> int:
    count = 0
    for mid in material_ids:
        material = db.query(RawMaterial).filter(RawMaterial.id == mid).first()
        if material:
            material.is_active = False
            count += 1
    db.commit()
    logger.info("Material bulk delete (deactivate): %d materials", count)
    return count
