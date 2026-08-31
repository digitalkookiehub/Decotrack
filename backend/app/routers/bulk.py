"""Bulk import/export endpoints for Vendors and Raw Materials."""
from typing import Annotated

from fastapi import APIRouter, Depends, File, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import require_any
from app.models.user import User
from app.services import bulk_service

router = APIRouter(prefix="/bulk", tags=["Bulk Import/Export"])


class BulkDeleteRequest(BaseModel):
    ids: list[int]


# ── Vendors ────────────────────────────────────────────────────

@router.get("/vendors/export")
async def export_vendors(
    db: Annotated[Session, Depends(get_db)],
    _user: Annotated[User, Depends(require_any)],
):
    output = bulk_service.export_vendors_excel(db)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vendors.xlsx"},
    )


@router.post("/vendors/import")
async def import_vendors(
    db: Annotated[Session, Depends(get_db)],
    _user: Annotated[User, Depends(require_any)],
    file: UploadFile = File(...),
):
    contents = await file.read()
    result = bulk_service.import_vendors_excel(db, contents)
    return result


@router.post("/vendors/delete")
async def delete_vendors_bulk(
    data: BulkDeleteRequest,
    db: Annotated[Session, Depends(get_db)],
    _user: Annotated[User, Depends(require_any)],
):
    count = bulk_service.delete_vendors_bulk(db, data.ids)
    return {"deleted": count}


# ── Raw Materials ──────────────────────────────────────────────

@router.get("/materials/export")
async def export_materials(
    db: Annotated[Session, Depends(get_db)],
    _user: Annotated[User, Depends(require_any)],
):
    output = bulk_service.export_materials_excel(db)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=raw_materials.xlsx"},
    )


@router.post("/materials/import")
async def import_materials(
    db: Annotated[Session, Depends(get_db)],
    _user: Annotated[User, Depends(require_any)],
    file: UploadFile = File(...),
):
    contents = await file.read()
    result = bulk_service.import_materials_excel(db, contents)
    return result


@router.post("/materials/delete")
async def delete_materials_bulk(
    data: BulkDeleteRequest,
    db: Annotated[Session, Depends(get_db)],
    _user: Annotated[User, Depends(require_any)],
):
    count = bulk_service.delete_materials_bulk(db, data.ids)
    return {"deleted": count}


# ── Template Downloads ─────────────────────────────────────────

@router.get("/vendors/template")
async def vendor_template():
    from openpyxl import Workbook
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"
    ws.append(["ID (leave blank for new)", "Name*", "Contact Person", "Phone", "Email", "Address", "City", "State", "GSTIN", "Supply Categories (comma-sep)", "Payment Terms", "Notes"])
    ws.append([None, "Example Vendor", "John", "9876543210", "john@vendor.com", "123 Street", "Chennai", "Tamil Nadu", "33AABCU9603R1ZM", "Sheet Materials, Hardware", "Net 30", "Good supplier"])

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=vendor_template.xlsx"})


@router.get("/materials/template")
async def material_template():
    from openpyxl import Workbook
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Materials"
    ws.append(["ID (leave blank for new)", "Name*", "SKU*", "Category", "Unit* (PCS/SQM/M/KG/PAIR/LITRE/SET)", "Current Stock", "Reorder Level*", "Last Purchase Rate", "HSN Code", "GST Rate (%)", "Description"])
    ws.append([None, "Plywood 18mm BWR", "PLY-18-BWR", "Sheet Materials", "SQM", 50, 10, 95, "4412", 18, "18mm BWR grade plywood"])

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=material_template.xlsx"})
